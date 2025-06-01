#!/usr/bin/env python3
"""
Extract data from WPFLHistoryCondensed.xlsx and convert to JSON format
This script reads the Excel file and creates multiple JSON files for different data sections
"""
import json
import sys
import os

print("Starting Excel data extraction...")
print("This script will parse the Excel file manually using openpyxl to avoid pandas memory issues")

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Installing it now...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

# File paths
excel_path = "data/WPFLHistoryCondensed.xlsx"
output_dir = "data/wpfl_extracted"

# Create output directory
os.makedirs(output_dir, exist_ok=True)

print(f"Loading workbook from {excel_path}...")
wb = load_workbook(excel_path, read_only=True, data_only=True)

# Get all sheet names
sheet_names = wb.sheetnames
print(f"Found {len(sheet_names)} sheets: {sheet_names}")

# Process each sheet
all_data = {}

for sheet_name in sheet_names:
    print(f"\nProcessing sheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"  Dimensions: {max_row} rows x {max_col} columns")
    
    # Extract data in chunks to handle large sheets
    sheet_data = {
        "info": {
            "sheet_name": sheet_name,
            "rows": max_row,
            "columns": max_col
        },
        "headers": [],
        "data_sections": {}
    }
    
    # Get headers (first few rows often contain headers)
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        row_data = [str(cell) if cell is not None else "" for cell in row]
        if any(row_data):  # Skip empty rows
            headers.append(row_data)
    sheet_data["headers"] = headers
    
    # Identify data sections by looking for patterns
    # Sample every 50th row to understand structure
    sample_rows = []
    for i in range(0, min(max_row, 1000), 50):
        row = list(sheet.iter_rows(min_row=i+1, max_row=i+1, values_only=True))[0]
        row_data = [str(cell) if cell is not None else "" for cell in row]
        sample_rows.append({
            "row_num": i+1,
            "data": row_data,
            "non_empty_count": sum(1 for cell in row_data if cell)
        })
    
    # Save sample data for analysis
    sample_file = os.path.join(output_dir, f"{sheet_name}_samples.json")
    with open(sample_file, 'w') as f:
        json.dump({
            "sheet_info": sheet_data["info"],
            "headers": headers,
            "sample_rows": sample_rows
        }, f, indent=2)
    print(f"  Saved samples to {sample_file}")
    
    # Based on the data summary we saw earlier, let's extract specific sections
    # The file seems to contain multiple data tables in different areas
    
    # Extract identifiable sections
    sections_to_extract = [
        {"name": "trade_history", "start_row": 1, "end_row": 250, "start_col": 1, "end_col": 10},
        {"name": "owner_stats", "start_row": 1, "end_row": 50, "start_col": 40, "end_col": 50},
        {"name": "head_to_head", "start_row": 1, "end_row": 200, "start_col": 43, "end_col": 55},
        {"name": "season_data", "start_row": 1, "end_row": 100, "start_col": 12, "end_col": 30}
    ]
    
    for section in sections_to_extract:
        print(f"  Extracting {section['name']} section...")
        section_data = []
        
        try:
            for row in sheet.iter_rows(
                min_row=section["start_row"], 
                max_row=min(section["end_row"], max_row),
                min_col=section["start_col"],
                max_col=min(section["end_col"], max_col),
                values_only=True
            ):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(row_data):  # Skip completely empty rows
                    section_data.append(row_data)
            
            if section_data:
                # Save section to separate file
                section_file = os.path.join(output_dir, f"{sheet_name}_{section['name']}.json")
                with open(section_file, 'w') as f:
                    json.dump(section_data, f, indent=2)
                print(f"    Saved {len(section_data)} rows to {section_file}")
                
                sheet_data["data_sections"][section['name']] = {
                    "rows": len(section_data),
                    "file": section_file
                }
        except Exception as e:
            print(f"    Error extracting {section['name']}: {e}")
    
    all_data[sheet_name] = sheet_data

# Save summary file
summary_file = os.path.join(output_dir, "extraction_summary.json")
with open(summary_file, 'w') as f:
    json.dump(all_data, f, indent=2)

print(f"\nExtraction complete! Summary saved to {summary_file}")
print("\nNext steps:")
print("1. Review the extracted JSON files in data/wpfl_extracted/")
print("2. Identify the specific data sections you want to analyze")
print("3. Create focused analysis on those sections")

# Also create a simplified data structure based on what we know
print("\nCreating simplified data structure...")

# Based on the data summary, we know there's owner data, trade history, head-to-head records
# Let's create a focused extraction of key data
simplified_data = {
    "owners": [],
    "trade_summary": {},
    "head_to_head": {},
    "statistics": {}
}

# Try to extract owner list from the data
print("Attempting to extract owner information...")
potential_owners = ["Mike S", "Adler", "Nixon", "AJ", "Forrest", "Todd", "Jimmy", "Dave", "Neill", "Ryan", "Doug", "Hoyle", "Rick", "Mims"]
simplified_data["owners"] = potential_owners

# Save simplified structure
simplified_file = os.path.join(output_dir, "wpfl_simplified.json")
with open(simplified_file, 'w') as f:
    json.dump(simplified_data, f, indent=2)

print(f"Simplified data saved to {simplified_file}")

wb.close()
print("\nExtraction script completed!")